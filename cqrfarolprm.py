# Salve este código como: maestro_app.py

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
from datetime import datetime
import pyodbc
import io
import random
from openpyxl.styles import Font, PatternFill, Alignment

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO DA PÁGINA E CSS PREMIUM
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="MAESTRO FAROL", page_icon="🔮", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    .stApp { background-color: #000000; color: #E0E0E0; }
    .header-premium { background: linear-gradient(135deg, #1e2a52 0%, #3b1d5a 100%); padding: 25px; border-radius: 15px; margin-bottom: 25px; border: 1px solid rgba(79, 195, 247, 0.3); box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37); }
    .logo-maestro { font-size: 2.5em; font-weight: bold; color: #ffffff; text-shadow: 0 0 10px rgba(79, 195, 247, 0.7); }
    .tagline { color: #b0c4de; font-style: italic; font-size: 1.1em; }
    .ceo-dashboard { background: #0a0a0a; backdrop-filter: blur(10px); border-radius: 15px; padding: 25px; margin-bottom: 20px; border: 1px solid rgba(79, 195, 247, 0.2); box-shadow: 0 0 40px rgba(79, 195, 247, 0.1), inset 0 0 15px rgba(0,0,0,0.5); }
    .prescription-card { background: rgba(255, 255, 255, 0.05); border-left: 5px solid; border-radius: 8px; padding: 20px; margin-bottom: 20px; transition: all 0.3s ease; }
    .prescription-card.CRÍTICA { border-left-color: #FF4500; } .prescription-card.ALTA { border-left-color: #FFA500; }
    .prescription-card.MÉDIA { border-left-color: #4FC3F7; } .prescription-card.BAIXA { border-left-color: #4CAF50; }
    .prescription-title { color: #ffffff; margin: 10px 0 5px 0; font-weight: bold; }
    .prescription-icon { font-size: 1.5em; margin-right: 10px; }
    .socratic-card { background: rgba(255, 215, 0, 0.05); border: 1px solid rgba(255, 215, 0, 0.2); border-left: 5px solid #FFD700; border-radius: 8px; padding: 25px; margin-bottom: 20px; }
    .socratic-title { color: #FFD700; font-weight: bold; font-size: 1.2em; }
    blockquote { border-left: 3px solid #b0c4de; padding-left: 1.5rem; margin-left: 1rem; font-style: italic; color: #E0E0E0; font-size: 1.1em; }
    .voice-section { background: #0a0a0a; border-radius: 15px; padding: 25px; margin: 20px 0; border: 1px solid rgba(79, 195, 247, 0.2); text-align: center; }
    [data-testid="stMetricValue"] { color: #4FC3F7; }
    [data-testid="stMetricDelta"] svg { display: inline; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO DE DADOS E FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=300)
def carregar_universo_de_dados():
    try:
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={st.secrets['db_credentials']['server']};"
            f"DATABASE={st.secrets['db_credentials']['database']};"
            f"UID={st.secrets['db_credentials']['username']};"
            f"PWD={st.secrets['db_credentials']['password']};"
            f"TrustServerCertificate=yes;"
        )
        cnxn = pyodbc.connect(conn_str)
        
        super_query = """
        SELECT
            CAST(g.Mes as INT) as Mes, CAST(g.Ano as INT) as Ano,
            g.QtHrOrc as "Hrs_Prev", g.QtHrReal as "Hrs_Real",
            g.ReceitaReal as "Receita", g.CustoReal as "Custo",
            tec.NomeTec as "Consultor", niv.DescNivel as "Nivel_Consultor",
            p.DescProj as "Projeto", t.DescTipo as "TipoProj",
            cli.DescCli as "Cliente", p.VlHrProj as "VH Venda",
            tec.VlHrTec as "VH Custo"
        FROM Tb_GestorFin2 g
        LEFT JOIN tb_Proj p ON g.ProjGest = p.AutNumProj
        LEFT JOIN tb_tec tec ON g.ConsultGest = tec.AutNumTec
        LEFT JOIN tb_cli cli ON p.CodCliProj = cli.AutNumCli
        LEFT JOIN tb_tipoproj t ON p.TipoProj = t.AutNumTipo
        LEFT JOIN tb_amarradisc amarra ON tec.AutNumTec = amarra.CodTecAmar
        LEFT JOIN tb_nivel niv ON amarra.Nivel = niv.AutNivel -- CORREÇÃO DO ERRO
        WHERE tec.NomeTec IS NOT NULL AND p.DescProj IS NOT NULL
        GROUP BY
            g.IdGest2, g.Mes, g.Ano, g.QtHrOrc, g.QtHrReal, g.ReceitaReal, g.CustoReal,
            g.PercMgReal, tec.NomeTec, niv.DescNivel, p.DescProj, t.DescTipo, cli.DescCli,
            p.VlHrProj, tec.VlHrTec;
        """
        df = pd.read_sql(super_query, cnxn)
        cnxn.close()

        for col in ["Hrs_Prev", "Hrs_Real", "Receita", "Custo", "VH Venda", "VH Custo"]:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        df['Dt_Fim_Real'] = pd.to_datetime(df['Ano'].astype(str) + '-' + df['Mes'].astype(str) + '-01', errors='coerce')
        df['TipoContrato'] = df['TipoProj'].apply(lambda x: 'Escopo Fechado' if 'Fechado' in str(x) else 'Time & Material')
        df['Receita_Esperada'] = df['Hrs_Real'] * df['VH Venda']
        df['Valor_Nao_Faturado'] = df['Receita_Esperada'] - df['Receita']
        df['Valor_Nao_Faturado'] = df['Valor_Nao_Faturado'].apply(lambda x: x if x > 0 else 0)
        df['Lucro'] = df['Receita'] - df['Custo']
        df['Margem'] = np.where(df['Receita'] > 0, (df['Lucro'] / df['Receita']) * 100, 0)
        df['Desvio_Hrs'] = df['Hrs_Real'] - df['Hrs_Prev']
        
        return df.fillna(0)
    except Exception as e:
        st.error(f"Erro Crítico ao conectar com o universo de dados: {e}")
        return pd.DataFrame()

def to_excel_formatted(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Fechamento')
        workbook = writer.book
        worksheet = writer.sheets['Fechamento']
        header_fill = PatternFill(start_color="1E2A52", end_color="1E2A52", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            worksheet.column_dimensions[cell.column_letter].width = 18
    return output.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# NÚCLEO DE RACIOCÍNIO QUÂNTICO (CRQ) - VERSÃO FINAL COMPLETA
# ═══════════════════════════════════════════════════════════════════════════════
class CoreQuantumReasoning:
    def __init__(self, dados_universo):
        self.dados_universo = dados_universo
        self.estado_quantum = pd.DataFrame()

    def aplicar_colapso_quantico(self, filtros):
        df = self.dados_universo.copy()
        if filtros.get('mes') and 'TODOS' not in str(filtros['mes']): df = df[df['Mes'] == filtros['mes']]
        if filtros.get('ano') and 'TODOS' not in str(filtros['ano']): df = df[df['Ano'] == filtros['ano']]
        if filtros.get('consultores') and 'TODOS' not in filtros['consultores']: df = df[df['Consultor'].isin(filtros['consultores'])]
        if filtros.get('clientes') and 'TODOS' not in filtros['clientes']: df = df[df['Cliente'].isin(filtros['clientes'])]
        self.estado_quantum = df

    def gerar_sinfonia_de_insights_sapiens(self, filtros):
        df = self.estado_quantum
        if df.empty: return []
        insights = []
        periodo_str = "no período analisado"

        # INSIGHT 1: Vazamento de Lucro em Projetos Fechados
        proj_fechados_overrun = df[(df['TipoContrato'] == 'Escopo Fechado') & (df['Desvio_Hrs'] > 0)]
        if not proj_fechados_overrun.empty:
            custo_extra = (proj_fechados_overrun['Desvio_Hrs'] * proj_fechados_overrun['VH Custo']).sum()
            margem_media = proj_fechados_overrun['Margem'].mean()
            if margem_media < 25 and custo_extra > 500:
                insights.append({'tipo': 'FINANCEIRO', 'prioridade': 'CRÍTICA', 'icone': '🚱', 'titulo': 'Vazamento de Lucro em Projetos Fechados', 'analise': f"O excesso de horas em projetos de 'Escopo Fechado' consumiu R$ {custo_extra:,.2f} do lucro {periodo_str}, derrubando a margem média destes para {margem_media:.1f}%.", 'prescricao': "1. **Contenção:** Revisar imediatamente o escopo e o progresso dos projetos afetados.\n2. **Prevenção:** Implementar um processo rigoroso de 'Change Request'.\n3. **Estratégia:** Calibrar futuras propostas com uma margem de segurança de 15-20% nas horas."})

        # INSIGHT 2: Receita Não Realizada (Horas Não Faturadas)
        total_nao_faturado = df['Valor_Nao_Faturado'].sum()
        if not df.empty and df['Receita'].sum() > 0 and total_nao_faturado > (df['Receita'].sum() * 0.05):
            insights.append({'tipo': 'FINANCEIRO', 'prioridade': 'ALTA', 'icone': '🧾', 'titulo': 'Receita Não Realizada (Horas Não Faturadas)', 'analise': f"Detectamos um gap de R$ {total_nao_faturado:,.2f} entre as horas trabalhadas e o valor faturado {periodo_str}. Isso pode indicar falhas no processo de faturamento ou 'cortesias' não documentadas.", 'prescricao': "1. **Auditoria:** Realizar conciliação focada nos projetos com gap.\n2. **Processo:** Garantir que 100% das horas em projetos 'Time & Material' sejam refletidas na fatura.\n3. **Política:** Definir e registrar políticas de desconto ou investimento."})

        # INSIGHT 3: Ponto de Atenção Estratégica em T&M
        proj_tm_overrun = df[(df['TipoContrato'] == 'Time & Material') & (df['Desvio_Hrs'] > 0) & (df['Valor_Nao_Faturado'] == 0)]
        if not proj_tm_overrun.empty:
            receita_adicional = (proj_tm_overrun['Desvio_Hrs'] * proj_tm_overrun['VH Venda']).sum()
            if not df.empty and df['Receita'].sum() > 0 and receita_adicional > (df['Receita'].sum() * 0.1):
                insights.append({'tipo': 'ESTRATÉGICO', 'prioridade': 'MÉDIA', 'icone': '🧐', 'titulo': 'Atenção: Expansão de Escopo Validada', 'analise': f"**Validação:** O sistema confirmou que o aumento de horas em {len(proj_tm_overrun)} projetos 'Time & Material' foi corretamente faturado, gerando R$ {receita_adicional:,.2f} de receita adicional. Financeiramente, está correto.\n\n**Ponto de Atenção:** Este padrão indica uma expansão contínua do escopo ('scope creep'). Se não for gerenciado, pode levar a desalinhamentos de expectativa e impactar a alocação de recursos.", 'prescricao': "1. **Alinhamento:** Agendar reunião com o cliente para discutir o roadmap e formalizar a expansão.\n2. **Gestão de Recursos:** Avaliar se a equipe alocada ainda é suficiente para a nova dimensão do projeto.\n3. **Oportunidade de Upsell:** Transformar a necessidade crescente em um novo contrato."})
        
        # INSIGHT 4: Assimetria de Performance entre Consultores
        if df['Consultor'].nunique() > 2:
            perf_cons = df.groupby('Consultor').agg(Lucro_Gerado=('Lucro', 'sum')).sort_values('Lucro_Gerado', ascending=False)
            top_performer, lucro_top = perf_cons.index[0], perf_cons.iloc[0]['Lucro_Gerado']
            bottom_performer, lucro_bottom = perf_cons.index[-1], perf_cons.iloc[-1]['Lucro_Gerado']
            if lucro_top > 0 and (lucro_top / max(abs(lucro_bottom), 1)) > 3 :
                 insights.append({'tipo': 'TALENTO', 'prioridade': 'ALTA', 'icone': '🏆', 'titulo': 'Assimetria de Performance Detectada', 'analise': f"Existe uma disparidade significativa na geração de lucro {periodo_str}. O consultor **{top_performer}** gerou R$ {lucro_top:,.2f} em lucro, enquanto **{bottom_performer}** teve um resultado de R$ {lucro_bottom:,.2f}. Nivelar essa performance é uma alavanca de crescimento.", 'prescricao': f"1. **Mentoria:** Implementar programa de mentoria: {top_performer} → {bottom_performer}.\n2. **Padronização:** Documentar e disseminar as melhores práticas do Top Performer.\n3. **Alocação:** Avaliar se os projetos de menor resultado estão adequados ao nível do consultor."})

        # INSIGHT 5: Risco de Burnout (Com Lógica de Filtro Correta)
        if 'TODOS' not in str(filtros.get('mes')): # Só roda se um mês específico for selecionado
            carga_consultor = df.groupby('Consultor')['Hrs_Real'].sum()
            if not carga_consultor.empty:
                limite_horas_mes = 180 
                consultores_sobrecarregados = carga_consultor[carga_consultor > limite_horas_mes]
                if not consultores_sobrecarregados.empty:
                    cons_critico, horas_criticas = consultores_sobrecarregados.idxmax(), consultores_sobrecarregados.max()
                    insights.append({'tipo': 'CAPITAL HUMANO', 'prioridade': 'ALTA', 'icone': '🔥', 'titulo': 'Risco de Saturação de Equipe', 'analise': f"No mês selecionado, o consultor **{cons_critico}** registrou **{horas_criticas:.0f} horas**. Este volume é insustentável e representa um risco à qualidade e bem-estar.", 'prescricao': "1. **Diálogo Preventivo:** Validar com o consultor se o volume de horas reflete o esforço real.\n2. **Análise Qualitativa:** A carga de trabalho é de alta complexidade? Exige deslocamento?\n3. **Ação de Balanceamento:** Planejar o próximo ciclo com uma alocação mais leve."})

        if not insights:
             insights.append({'tipo': 'SUCESSO', 'prioridade': 'BAIXA', 'icone': '✅', 'titulo': 'Operação Consistente', 'analise': f"Nenhuma dissonância crítica foi detectada {periodo_str}. Os processos financeiros e operacionais estão funcionando como esperado.", 'prescricao': "Manter a disciplina e os controles atuais."})
        
        return sorted(insights, key=lambda p: ['CRÍTICA', 'ALTA', 'MÉDIA', 'BAIXA'].index(p['prioridade']))

    def gerar_perguntas_socraticas(self):
        df = self.estado_quantum
        if df.empty or len(df) < 2: return []
        perguntas = []

        if df['TipoProj'].nunique() > 1 and df['Receita'].sum() > 0:
            rentabilidade = df.groupby('TipoProj').agg(Receita=('Receita', 'sum'), Margem=('Margem', 'mean')).reset_index()
            rentabilidade = rentabilidade[(rentabilidade['Receita'] > 0) & (rentabilidade['Margem'] > 0)]
            if len(rentabilidade) > 1:
                maior_receita = rentabilidade.loc[rentabilidade['Receita'].idxmax()]
                maior_margem = rentabilidade.loc[rentabilidade['Margem'].idxmax()]
                if maior_receita['TipoProj'] != maior_margem['TipoProj'] and (maior_margem['Margem'] - maior_receita['Margem']) > 10:
                    perguntas.append({'icone': '🤔', 'titulo': 'Sobre a Estratégia de Mix de Serviços', 'pergunta': f"O Maestro observa que '{maior_receita['TipoProj']}' gerou {(maior_receita['Receita'] / df['Receita'].sum() * 100):.0f}% da receita com uma margem de {maior_receita['Margem']:.1f}%, enquanto '{maior_margem['TipoProj']}' é {(maior_margem['Margem'] / maior_receita['Margem']):.1f}x mais lucrativo. Esta é uma alavancagem estratégica para ganhar mercado, ou existe uma oportunidade para otimizar o nosso mix de serviços?"})
        
        is_saudavel = 'Vazamento de Lucro' not in [i['titulo'] for i in self.gerar_sinfonia_de_insights_sapiens({})]
        if is_saudavel and not df.empty and df['Margem'].mean() > 35:
            perguntas.append({'icone': '🚀', 'titulo': 'Sobre o Próximo Horizonte de Crescimento', 'pergunta': "A operação demonstra uma saúde excepcional. Com esta base sólida, qual seria o próximo 'salto quântico': expandir para um novo tipo de serviço, ou aprofundar a rentabilidade nos clientes mais estratégicos que já possuímos?"})

        return perguntas
    
    def analisar_fluxo_de_caixa(self):
        df = self.estado_quantum
        if df.empty: return {'dados': [], 'analise': "Sem dados para análise de fluxo de caixa."}
        df_fluxo = df.copy()
        df_fluxo['Periodo'] = pd.to_datetime(df_fluxo['Ano'].astype(str) + '-' + df_fluxo['Mes'].astype(str) + '-01')
        fluxo = df_fluxo.groupby('Periodo').agg(Receber=('Receita', 'sum'), Pagar=('Custo', 'sum')).reset_index()
        fluxo['Saldo'] = fluxo['Receber'] - fluxo['Pagar']
        fluxo['name'] = fluxo['Periodo'].dt.strftime('%b/%Y')
        saldo_total = fluxo['Saldo'].sum()
        analise = f"O fluxo de caixa operacional do período é **{'POSITIVO' if saldo_total >= 0 else 'NEGATIVO'} em R$ {abs(saldo_total):,.2f}**."
        return {'dados': fluxo.to_dict('records'), 'analise': analise}

    def calcular_metricas_consolidadas(self):
        df = self.estado_quantum
        if df.empty: return {'receita': 0, 'lucro': 0, 'margem': 0, 'projetos': 0, 'consultores': 0, 'clientes': 0, 'hrs_real': 0, 'desvio_hrs': 0}
        return {
            'receita': df['Receita'].sum(), 'lucro': df['Lucro'].sum(), 'margem': df['Margem'].mean(),
            'projetos': df['Projeto'].nunique(), 'consultores': df['Consultor'].nunique(),
            'clientes': df['Cliente'].nunique(), 'hrs_real': df['Hrs_Real'].sum(), 'desvio_hrs': df['Desvio_Hrs'].sum()
        }

# ═══════════════════════════════════════════════════════════════════════════════
# ORQUESTRAÇÃO PRINCIPAL E INTERFACE STREAMLIT
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="header-premium"><div class="logo-maestro">🔮 MAESTRO FAROL</div><div class="tagline">A Orquestra de Realidades para a Gestão de Negócios</div></div>', unsafe_allow_html=True)

dados_universo = carregar_universo_de_dados()

if not dados_universo.empty:
    crq = CoreQuantumReasoning(dados_universo)

    with st.sidebar:
        st.markdown("### 🧭 Controles do Universo")
        anos_disponiveis = ["TODOS"] + sorted(dados_universo['Ano'].unique().tolist(), reverse=True)
        meses_disponiveis = ["TODOS"] + list(range(1, 13))
        
        ano_sel = st.selectbox("Ano", anos_disponiveis, index=0)
        mes_sel = st.selectbox("Mês", meses_disponiveis, index=0)
        cons_sel = st.multiselect("Consultores", ["TODOS"] + sorted(dados_universo['Consultor'].unique().tolist()), default=["TODOS"])
        cli_sel = st.multiselect("Clientes", ["TODOS"] + sorted(dados_universo['Cliente'].unique().tolist()), default=["TODOS"])

    filtros = {'mes': mes_sel, 'ano': ano_sel, 'consultores': cons_sel, 'clientes': cli_sel}
    crq.aplicar_colapso_quantico(filtros)
    
    metricas = crq.calcular_metricas_consolidadas()
    insights = crq.gerar_sinfonia_de_insights_sapiens(filtros)
    perguntas = crq.gerar_perguntas_socraticas()
    fluxo_caixa = crq.analisar_fluxo_de_caixa()

    tab_exec, tab_ia, tab_soc, tab_fec, tab_fluxo, tab_comp, tab_cmd = st.tabs([
        "🎯 Visão Executiva", "🧠 Ressonância CQR", "❓ Perguntas Estratégicas", 
        "💰 Fechamento", "💸 Fluxo de Caixa", "📊 Comparativo", "🎤 Comandos"
    ])

    with tab_exec:
        st.markdown('<div class="ceo-dashboard">', unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("💰 Receita Total", f"R$ {metricas['receita']:,.0f}")
        c2.metric("📈 Lucro Líquido", f"R$ {metricas['lucro']:,.0f}")
        c3.metric("📊 Margem Média", f"{metricas['margem']:.1f}%")
        delta_hrs_color = "normal" if metricas['desvio_hrs'] <= 0 else "inverse"
        c4.metric("⏱️ Horas Realizadas", f"{metricas['hrs_real']:.0f}h", f"{metricas['desvio_hrs']:+.0f}h vs. Previsto", delta_color=delta_hrs_color)
        st.markdown("---")
        
        g1, g2 = st.columns(2)
        with g1:
            rec_por_cliente = crq.estado_quantum.groupby('Cliente')['Receita'].sum().sort_values(ascending=False).head(10)
            if not rec_por_cliente.empty:
                fig = px.bar(rec_por_cliente, x=rec_por_cliente.index, y=rec_por_cliente.values, title="Top 10 Clientes por Receita", template='plotly_dark', labels={'y': 'Receita', 'x': 'Cliente'})
                fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
                st.plotly_chart(fig, use_container_width=True)
        with g2:
            lucro_por_tipo = crq.estado_quantum.groupby('TipoProj')['Lucro'].sum().sort_values(ascending=False)
            if not lucro_por_tipo.empty:
                fig2 = px.pie(values=lucro_por_tipo.values, names=lucro_por_tipo.index, title="Contribuição para o Lucro por Tipo de Projeto", template='plotly_dark')
                fig2.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
                st.plotly_chart(fig2, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
    with tab_ia:
        if insights:
            for i in insights:
                st.markdown(f'<div class="prescription-card {i["prioridade"]}">', unsafe_allow_html=True)
                st.markdown(f'<h4 class="prescription-title"><span class="prescription-icon">{i["icone"]}</span> {i["titulo"]}</h4>', unsafe_allow_html=True)
                st.markdown(i['analise'].replace('\n', '<br>'), unsafe_allow_html=True)
                st.markdown(f"**Diretriz Recomendada:**\n{i['prescricao']}")
                st.markdown('</div>', unsafe_allow_html=True)

    with tab_soc:
        if perguntas:
            for p in perguntas:
                st.markdown('<div class="socratic-card">', unsafe_allow_html=True)
                st.markdown(f'<p class="socratic-title"><span class="prescription-icon">{p["icone"]}</span> {p["titulo"]}</p>', unsafe_allow_html=True)
                st.markdown(f"<blockquote>{p['pergunta']}</blockquote>")
                st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("No período atual, os dados são conclusivos e não levantaram dilemas estratégicos que exijam reflexão socrática.")

    with tab_fec:
        st.markdown("### 💰 Fechamento por Consultor")
        df_fec_cons = crq.estado_quantum.groupby('Consultor').agg(Receita=('Receita', 'sum'), Custo=('Custo', 'sum'), Lucro=('Lucro', 'sum'), Horas=('Hrs_Real', 'sum')).reset_index()
        st.dataframe(df_fec_cons.style.format({'Receita': 'R$ {:,.2f}', 'Custo': 'R$ {:,.2f}', 'Lucro': 'R$ {:,.2f}', 'Horas': '{:.1f}h'}), use_container_width=True)
        st.download_button("📥 Exportar Consultores (XLSX)", to_excel_formatted(df_fec_cons), "fechamento_consultores.xlsx")
        
        st.markdown("### 🏢 Fechamento por Cliente")
        df_fec_cli = crq.estado_quantum.groupby('Cliente').agg(Receita=('Receita', 'sum'), Custo=('Custo', 'sum'), Lucro=('Lucro', 'sum')).reset_index()
        st.dataframe(df_fec_cli.style.format({'Receita': 'R$ {:,.2f}', 'Custo': 'R$ {:,.2f}', 'Lucro': 'R$ {:,.2f}'}), use_container_width=True)
        st.download_button("📥 Exportar Clientes (XLSX)", to_excel_formatted(df_fec_cli), "fechamento_clientes.xlsx")

    with tab_fluxo:
        st.markdown(fluxo_caixa['analise'], unsafe_allow_html=True)
        if fluxo_caixa['dados']:
            fig = px.bar(fluxo_caixa['dados'], x='name', y=['Receber', 'Pagar'], 
                         title="A Receber (Receita) vs. A Pagar (Custo)", barmode='group',
                         labels={'name': ''}, template='plotly_dark', color_discrete_map={'Receber': '#4CAF50', 'Pagar': '#FF4500'})
            fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig, use_container_width=True)

    with tab_comp:
        st.markdown("### ⚖️ Comparativo de Realidades")
        c1, c2 = st.columns(2)
        anos_u = sorted(dados_universo['Ano'].unique())
        meses_u = sorted(dados_universo['Mes'].unique())
        with c1:
            st.markdown("##### Período 1")
            ano1 = st.selectbox("Ano 1", anos_u, key='ano1')
            mes1 = st.selectbox("Mês 1", meses_u, key='mes1')
        with c2:
            st.markdown("##### Período 2")
            ano2 = st.selectbox("Ano 2", anos_u, index=min(1, len(anos_u)-1), key='ano2')
            mes2 = st.selectbox("Mês 2", meses_u, index=min(1, len(meses_u)-1), key='mes2')
        
        if st.button("Comparar Realidades", type="primary", use_container_width=True):
            df1 = dados_universo[(dados_universo['Ano'] == ano1) & (dados_universo['Mes'] == mes1)]
            df2 = dados_universo[(dados_universo['Ano'] == ano2) & (dados_universo['Mes'] == mes2)]
            
            if not df1.empty and not df2.empty:
                m1_lucro, m1_receita = df1['Lucro'].sum(), df1['Receita'].sum()
                m2_lucro, m2_receita = df2['Lucro'].sum(), df2['Receita'].sum()
                delta_lucro, delta_receita = m2_lucro - m1_lucro, m2_receita - m1_receita
                
                st.markdown('<div class="ceo-dashboard">', unsafe_allow_html=True)
                cc1, cc2 = st.columns(2)
                cc1.metric("Variação da Receita", f"R$ {m2_receita:,.2f}", f"R$ {delta_receita:,.2f}")
                cc2.metric("Variação do Lucro", f"R$ {m2_lucro:,.2f}", f"R$ {delta_lucro:,.2f}")
                
                narrativa = f"Ao comparar {mes2}/{ano2} com {mes1}/{ano1}, observamos uma variação de **R$ {delta_lucro:,.2f}** no lucro. "
                lucro_cli1, lucro_cli2 = df1.groupby('Cliente')['Lucro'].sum(), df2.groupby('Cliente')['Lucro'].sum()
                df_comp = pd.concat([lucro_cli1, lucro_cli2], axis=1, keys=['P1', 'P2']).fillna(0)
                df_comp['Variacao'] = df_comp['P2'] - df_comp['P1']
                
                maior_ganho = df_comp[df_comp['Variacao'] > 0].sort_values('Variacao', ascending=False).head(1)
                maior_perda = df_comp[df_comp['Variacao'] < 0].sort_values('Variacao', ascending=True).head(1)
                
                if not maior_ganho.empty: narrativa += f"O principal vetor positivo foi o cliente **{maior_ganho.index[0]}**, que contribuiu com um aumento de **R$ {maior_ganho['Variacao'].iloc[0]:,.2f}**. "
                if not maior_perda.empty: narrativa += f"Por outro lado, o resultado foi pressionado pelo cliente **{maior_perda.index[0]}**, com uma queda de **R$ {abs(maior_perda['Variacao'].iloc[0]):,.2f}**. "
                st.info(narrativa)
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.warning("Dados não disponíveis para um ou ambos os períodos selecionados.")

    with tab_cmd:
        st.markdown('<div class="voice-section">', unsafe_allow_html=True)
        st.markdown("### 🎤 Simulador de Comandos de Voz")
        st.text_input("Digite um comando em linguagem natural:", placeholder="Ex: 'qual a receita do cliente X?'", key="voice_cmd")
        st.info("O motor de Processamento de Linguagem Natural está a ser calibrado.")
        st.markdown('</div>', unsafe_allow_html=True)

else:
    st.error("A sinfonia está em silêncio. Não foi possível carregar o universo de dados. Verifique as credenciais e a conexão com o banco.")

