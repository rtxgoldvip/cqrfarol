import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import warnings
import random
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(
    page_title="Maestro Farol - Quantum Intelligence",
    page_icon="🔮",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- ESTILOS CSS PREMIUM (AMOLED REFINEMENT) ---
st.markdown("""
<style>
/* ... (Base CSS mantida e refinada) ... */
.stApp {
    background-color: #000000; /* Fundo Super AMOLED */
}

/* ESTILO DO HEADER PREMIUM */
.header-premium {
    background: linear-gradient(135deg, #1e2a52 0%, #3b1d5a 100%);
    padding: 25px;
    border-radius: 15px;
    margin-bottom: 25px;
    border: 1px solid rgba(79, 195, 247, 0.3);
    box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
}
.logo-maestro {
    font-size: 2.5em;
    font-weight: bold;
    color: #ffffff;
    text-shadow: 0 0 10px rgba(79, 195, 247, 0.7);
}
.tagline {
    color: #b0c4de;
    font-style: italic;
    font-size: 1.1em;
}

/* PAINEL EXECUTIVO UNIFICADO (CEO Dashboard) */
.ceo-dashboard {
    background: #0a0a0a; /* Fundo ultra-escuro para contraste */
    backdrop-filter: blur(10px);
    border-radius: 15px;
    padding: 25px;
    margin-bottom: 20px;
    border: 1px solid rgba(79, 195, 247, 0.2);
    box-shadow: 0 0 40px rgba(79, 195, 247, 0.1), inset 0 0 15px rgba(0,0,0,0.5); /* Sombra externa e interna para profundidade */
}

/* ÍNDICE DE SAÚDE QUÂNTICA */
.quantum-health-card {
    background: rgba(0, 0, 0, 0.3);
    border-radius: 50%;
    width: 200px;
    height: 200px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
    border: 3px solid; /* Borda mais grossa */
    margin: auto;
    transition: all 0.5s ease;
}
.quantum-health-score { font-size: 3.5em; font-weight: bold; color: white; text-shadow: 0 0 15px; }
.quantum-health-title { color: #b0c4de; font-weight: bold; }
.status-saudavel { border-color: #4CAF50; box-shadow: 0 0 25px #4CAF50, inset 0 0 10px #4CAF50; }
.status-atencao { border-color: #FFD700; box-shadow: 0 0 25px #FFD700, inset 0 0 10px #FFD700; }
.status-critico { border-color: #FF4500; box-shadow: 0 0 25px #FF4500, inset 0 0 10px #FF4500; }
.quantum-health-score.status-saudavel { color: #4CAF50; text-shadow: 0 0 15px #4CAF50; }
.quantum-health-score.status-atencao { color: #FFD700; text-shadow: 0 0 15px #FFD700; }
.quantum-health-score.status-critico { color: #FF4500; text-shadow: 0 0 15px #FF4500; }

/* CARDS DE PRESCRIÇÃO */
.prescription-card {
    background: rgba(255, 255, 255, 0.05);
    border-left: 5px solid;
    border-radius: 8px;
    padding: 20px;
    margin-bottom: 20px;
    transition: all 0.3s ease;
}
.prescription-card:hover {
    transform: translateY(-5px);
    box-shadow: 0 10px 20px rgba(0,0,0,0.2);
}
.prescription-card.CRÍTICA { border-left-color: #FF4500; }
.prescription-card.ALTA { border-left-color: #FFA500; }
.prescription-card.MÉDIA { border-left-color: #4FC3F7; }
.prescription-card.BAIXA { border-left-color: #4CAF50; }
.prescription-title { color: #ffffff; margin: 10px 0 5px 0; font-weight: bold; }
.prescription-icon { font-size: 1.5em; margin-right: 10px; }


/* SIMULADOR QUÂNTICO */
.simulator-section {
    background: rgba(0, 0, 0, 0.2);
    padding: 25px;
    border-radius: 15px;
    border: 1px solid rgba(79, 195, 247, 0.2);
}
.simulator-results {
    background: #0a0a0a;
    padding: 20px;
    border-radius: 10px;
    margin-top: 20px;
    border: 1px solid rgba(79, 195, 247, 0.1);
}
</style>
""", unsafe_allow_html=True)


# --- DADOS E VARIÁVEIS GLOBAIS ---
consultores = ['RAFAEL OLIVEIRA', 'CLEBER NEVES', 'ADRIANO AFONSO','LEANDRO GONCALVES', 'VALDINER APARECIDO', 'THIAGO MILANÊS']
clientes = ['AUTOZONE', 'TOTVS NOROESTE', 'HYDAC', 'TBC','TOTVS IP', 'TOTVS PAULISTA', 'Investimento']
tipos_proj = ['Implantação ERP', 'Desenvolvimento Customizado', 'Suporte Contínuo', 'Consultoria Estratégica']
niveis = {'RAFAEL OLIVEIRA': 'SÊNIOR', 'CLEBER NEVES': 'PLENO','ADRIANO AFONSO': 'ESPECIALISTA', 'LEANDRO GONCALVES': 'PLENO','VALDINER APARECIDO': 'PLENO', 'THIAGO MILANÊS': 'SÊNIOR'}
complexidades = ['Baixa', 'Média', 'Alta', 'Crítica']
senioridade_exigida = ['JÚNIOR', 'PLENO', 'SÊNIOR', 'ESPECIALISTA']


# --- NÚCLEO DE RACIOCÍNIO QUÂNTICO (CRQ) - VERSÃO DIAMANTE ---
class CoreQuantumReasoning:
    def __init__(self):
        self.dados_universo = pd.DataFrame()
        self.estado_quantum = pd.DataFrame()

    def carregar_universo_dados(self):
        np.random.seed(42)
        num_registros = 150
        hoje = datetime.now()
        
        df = pd.DataFrame({
            'Consultor': np.random.choice(consultores, num_registros),
            'Cliente': np.random.choice(clientes, num_registros),
            'Projeto': [f'PROJ_{1000+i}' for i in range(num_registros)],
            'TipoProj': np.random.choice(tipos_proj, num_registros),
            'Dt_Inicio_Proj': [hoje - timedelta(days=random.randint(30, 365)) for _ in range(num_registros)],
            'Complexidade': np.random.choice(complexidades, num_registros, p=[0.2, 0.4, 0.3, 0.1]),
            'Senioridade_Exigida': np.random.choice(senioridade_exigida, num_registros, p=[0.1, 0.4, 0.4, 0.1]),
        })

        df['Nivel_Consultor'] = df['Consultor'].map(niveis)
        df['Duracao_Prev_Dias'] = [random.randint(30, 120) for _ in range(num_registros)]
        df['Dt_Fim_Prev'] = df.apply(lambda row: row['Dt_Inicio_Proj'] + timedelta(days=row['Duracao_Prev_Dias']), axis=1)
        
        atraso_real = np.random.normal(5, 20, num_registros).clip(-15, 60)
        df['Atraso_Dias'] = atraso_real.astype(int)
        df['Dt_Fim_Real'] = df.apply(lambda row: row['Dt_Fim_Prev'] + timedelta(days=row['Atraso_Dias']), axis=1)

        df['Hrs_Prev'] = df['Duracao_Prev_Dias'] * np.random.uniform(4, 6, num_registros)
        df['Hrs_Real'] = df['Hrs_Prev'] + (df['Atraso_Dias'] * np.random.uniform(4, 6, num_registros))
        
        df['VH_Venda_Base'] = df['Complexidade'].map({'Baixa': 100, 'Média': 130, 'Alta': 160, 'Crítica': 200})
        df['VH_Venda'] = df['VH_Venda_Base'] * np.random.uniform(0.95, 1.1, num_registros)
        df['VH_Custo_Base'] = df['Nivel_Consultor'].map({'PLENO': 70, 'SÊNIOR': 90, 'ESPECIALISTA': 120})
        df['VH_Custo'] = df['VH_Custo_Base'] * np.random.uniform(0.98, 1.05, num_registros)

        df['Receita'] = df['Hrs_Real'] * df['VH_Venda']
        df['Custo'] = df['Hrs_Real'] * df['VH_Custo']
        df['Lucro'] = df['Receita'] - df['Custo']
        df['Margem'] = np.where(df['Receita'] > 0, (df['Lucro'] / df['Receita']) * 100, 0)
        
        df['Data'] = df['Dt_Fim_Real']
        df['Ano'] = df['Data'].dt.year
        df['Mes'] = df['Data'].dt.month

        df['Eficiencia'] = np.where(df['Hrs_Prev'] > 0, (df['Hrs_Real'] / df['Hrs_Prev']) * 100, 100)
        df['ROI_Hora'] = np.where(df['Hrs_Real'] > 0, (df['Receita'] - df['Custo']) / df['Hrs_Real'], 0)

        map_senioridade = {'JÚNIOR': 1, 'PLENO': 2, 'SÊNIOR': 3, 'ESPECIALISTA': 4}
        df['Num_Senioridade_Exigida'] = df['Senioridade_Exigida'].map(map_senioridade)
        df['Num_Nivel_Consultor'] = df['Nivel_Consultor'].map(map_senioridade)
        df['Mismatch_Senioridade'] = df['Num_Nivel_Consultor'] - df['Num_Senioridade_Exigida']
        
        df['Score_Risco'] = (
            (np.clip(df['Atraso_Dias'], 0, 60) / 60 * 0.5) +
            (np.clip(-df['Mismatch_Senioridade'], 0, 3) / 3 * 0.5)
        ) * 100
        
        df['Score_Performance'] = (
            (np.clip(df['Margem'], 0, 100) / 100 * 0.4) +
            (np.clip(100 - abs(df['Eficiencia'] - 100), 0, 100) / 100 * 0.2) +
            (np.clip(df['ROI_Hora'] / (df['VH_Venda'] - df['VH_Custo']).mean(), 0, 1.5) / 1.5 * 0.2) +
            ((100 - df['Score_Risco']) / 100 * 0.2)
        ) * 100
        
        self.dados_universo = df.fillna(0)

    def aplicar_colapso_quantico(self, filtros):
        df = self.dados_universo.copy()
        if filtros.get('consultores') and 'TODOS' not in filtros['consultores']:
            df = df[df['Consultor'].isin(filtros['consultores'])]
        if filtros.get('clientes') and 'TODOS' not in filtros['clientes']:
            df = df[df['Cliente'].isin(filtros['clientes'])]
        if filtros.get('tipos') and 'TODOS' not in filtros['tipos']:
            df = df[df['TipoProj'].isin(filtros['tipos'])]
        if filtros.get('mes') and 'TODOS' not in filtros['mes']:
            df = df[df['Mes'] == filtros['mes']]
        if filtros.get('ano') and 'TODOS' not in filtros['ano']:
            df = df[df['Ano'] == filtros['ano']]
        self.estado_quantum = df
        return df

    def calcular_indice_saude_quantica(self):
        df = self.estado_quantum
        if df.empty:
            return 0, "INDETERMINADO", "Sem dados para análise."

        margem_media = df['Margem'].mean()
        eficiencia_media = df['Eficiencia'].mean()
        atraso_medio = df['Atraso_Dias'].mean()
        mismatch_medio = df['Mismatch_Senioridade'].mean()

        score_margem = np.clip(margem_media / 35, 0, 1.5)
        score_eficiencia = np.clip(1 - abs(eficiencia_media - 100) / 50, 0, 1)
        score_prazo = np.clip(1 - atraso_medio / 30, 0, 1)
        score_alocacao = np.clip(1 - abs(mismatch_medio) / 1.5, 0, 1)
        
        peso_margem = 0.4; peso_eficiencia = 0.2; peso_prazo = 0.25; peso_alocacao = 0.15

        indice_final = (score_margem * peso_margem + score_eficiencia * peso_eficiencia + score_prazo * peso_prazo + score_alocacao * peso_alocacao) * 100
        indice_final = np.clip(indice_final, 0, 100)

        if indice_final >= 75:
            status, descricao = "SAUDÁVEL", "Ressonância positiva. Operação harmônica e lucrativa."
        elif indice_final >= 50:
            status, descricao = "ATENÇÃO", "Dissonância moderada. Pontos de melhoria detectados."
        else:
            status, descricao = "CRÍTICO", "Risco de colapso. Ações corretivas urgentes são necessárias."
        return int(indice_final), status, descricao

    def gerar_mapa_entrelacamento(self):
        df = self.estado_quantum
        if df.empty or len(df) < 2: return None
        cols_interesse = ['Margem', 'Atraso_Dias', 'Eficiencia', 'Mismatch_Senioridade', 'ROI_Hora']
        corr_matrix = df[cols_interesse].corr()
        fig = go.Figure(data=go.Heatmap(z=corr_matrix, x=corr_matrix.columns, y=corr_matrix.columns, hoverongaps=False, colorscale='RdBu_r', zmin=-1, zmax=1))
        fig.update_layout(title='Mapa de Entrelaçamentos (Correlações)', template='plotly_dark', paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        return fig

    def simular_cenario(self, df_base, alteracoes):
        df_simulado = df_base.copy()
        
        if alteracoes['tipo'] == 'custo':
            df_simulado['VH_Custo'] *= (1 + alteracoes['variacao'] / 100)

        elif alteracoes['tipo'] == 'realocacao' and alteracoes['pior_negocio'] and alteracoes['melhor_negocio'] != alteracoes['pior_negocio']:
            pior_negocio = alteracoes['pior_negocio']
            melhor_negocio = alteracoes['melhor_negocio']
            realoc_pct = alteracoes['percentual'] / 100

            horas_a_realocar = df_simulado.loc[df_simulado['TipoProj'] == pior_negocio, 'Hrs_Real'].sum() * realoc_pct
            if horas_a_realocar == 0: return df_simulado, False

            # Fatores de conversão (médias do melhor negócio)
            melhor_negocio_df = df_simulado[df_simulado['TipoProj'] == melhor_negocio]
            if melhor_negocio_df.empty or melhor_negocio_df['Hrs_Real'].sum() == 0: return df_simulado, False
            
            vh_venda_melhor = melhor_negocio_df['Receita'].sum() / melhor_negocio_df['Hrs_Real'].sum()
            vh_custo_melhor = melhor_negocio_df['Custo'].sum() / melhor_negocio_df['Hrs_Real'].sum()
            
            # Reduzir do pior negócio
            df_simulado.loc[df_simulado['TipoProj'] == pior_negocio, ['Receita', 'Custo', 'Hrs_Real']] *= (1 - realoc_pct)
            
            # Adicionar ao melhor negócio (de forma distribuída)
            receita_adicional = horas_a_realocar * vh_venda_melhor
            custo_adicional = horas_a_realocar * vh_custo_melhor
            
            indices_melhor = df_simulado.index[df_simulado['TipoProj'] == melhor_negocio]
            if not indices_melhor.empty:
                df_simulado.loc[indices_melhor, 'Receita'] += receita_adicional / len(indices_melhor)
                df_simulado.loc[indices_melhor, 'Custo'] += custo_adicional / len(indices_melhor)
                df_simulado.loc[indices_melhor, 'Hrs_Real'] += horas_a_realocar / len(indices_melhor)

        # Recalcula as métricas chave
        df_simulado['Lucro'] = df_simulado['Receita'] - df_simulado['Custo']
        df_simulado['Margem'] = np.where(df_simulado['Receita'] > 0, (df_simulado['Lucro'] / df_simulado['Receita']) * 100, 0)
        df_simulado['Atraso_Dias'] = df_simulado['Atraso_Dias'] # Placeholder para futuras simulações de atraso
        return df_simulado, True


    def gerar_prescricoes_quantum_premium(self, filtros):
        df = self.estado_quantum
        if df.empty: return []
        prescricoes = []

        # PRESCRIÇÃO 1: Dissonância de Alocação
        projetos_mismatch = df[df['Mismatch_Senioridade'] < -1]
        if not projetos_mismatch.empty:
            proj_critico = projetos_mismatch.loc[projetos_mismatch['Score_Risco'].idxmax()]
            prescricoes.append({'icone': '👥', 'tipo': 'ALOCAÇÃO', 'prioridade': 'CRÍTICA', 'titulo': 'Dissonância Crítica de Alocação',
                                'analise': f"O consultor {proj_critico['Consultor']} ({proj_critico['Nivel_Consultor']}) está em um projeto de alta complexidade ('{proj_critico['Projeto']}') que exige nível {proj_critico['Senioridade_Exigida']}. Este desalinhamento gera um risco sistêmico de atraso ({proj_critico['Atraso_Dias']:.0f} dias) e compromete a margem.",
                                'prescricao': "1. Ação Imediata: Realocar um consultor Sênior/Especialista para este projeto. \n2. Ação Estratégica: Revisar o processo de alocação para cruzar 'Complexidade do Projeto' com 'Nível do Consultor'."})

        # PRESCRIÇÃO 2: Projetos em Zona de Colapso
        projetos_colapso = df[(df['Margem'] < 15) & (df['Atraso_Dias'] > 20)]
        if not projetos_colapso.empty:
            receita_em_risco = projetos_colapso['Receita'].sum()
            prescricoes.append({'icone': '💥', 'tipo': 'RENTABILIDADE', 'prioridade': 'CRÍTICA', 'titulo': 'Projetos em Rota de Colapso Financeiro',
                                'analise': f"Detectamos {len(projetos_colapso)} projetos operando com margem crítica e atrasos significativos. Isso representa R$ {receita_em_risco:,.2f} em receita que está destruindo valor.",
                                'prescricao': "1. Comitê de Crise: Analisar individualmente cada um desses projetos. \n2. Renegociação: Iniciar renegociação de escopo/prazo com os clientes envolvidos. \n3. Controle de Danos: Avaliar a viabilidade de pausar ou encerrar projetos irrecuperáveis."})

        # PRESCRIÇÃO 3: Oportunidade de Ouro
        rentabilidade_tipo = df.groupby('TipoProj').agg({'ROI_Hora': 'mean', 'Hrs_Real': 'sum'}).reset_index()
        if len(rentabilidade_tipo) > 1:
            melhor_tipo = rentabilidade_tipo.loc[rentabilidade_tipo['ROI_Hora'].idxmax()]
            pior_tipo = rentabilidade_tipo.loc[rentabilidade_tipo['ROI_Hora'].idxmin()]
            if pior_tipo['ROI_Hora'] > 0 and (melhor_tipo['ROI_Hora'] / pior_tipo['ROI_Hora']) > 2:
                prescricoes.append({'icone': '💎', 'tipo': 'ESTRATÉGIA', 'prioridade': 'ALTA', 'titulo': 'Oportunidade de Ouro: Otimização do Mix de Serviços',
                                    'analise': f"O serviço de '{melhor_tipo['TipoProj']}' gera um ROI/Hora 2x maior que '{pior_tipo['TipoProj']}'. Atualmente, {(pior_tipo['Hrs_Real']/df['Hrs_Real'].sum()*100):.0f}% das horas estão alocadas no serviço menos rentável.",
                                    'prescricao': "1. Foco Comercial: Direcionar a força de vendas para priorizar contratos de '{melhor_tipo['TipoProj']}'. \n2. Repricing: Revisar a precificação dos serviços de '{pior_tipo['TipoProj']}'. \n3. Upselling: Criar pacotes para migrar clientes para serviços de maior valor."})

        # PRESCRIÇÃO 4: Risco de Burnout (REFINADO)
        if filtros.get('mes') and filtros.get('mes') != 'TODOS':
            carga_consultor = df.groupby('Consultor').agg({'Hrs_Real': 'sum'}).reset_index()
            consultores_sobrecarregados = carga_consultor[carga_consultor['Hrs_Real'] > 180]
            if not consultores_sobrecarregados.empty:
                cons_critico = consultores_sobrecarregados.loc[consultores_sobrecarregados['Hrs_Real'].idxmax()]
                prescricoes.append({'icone': '🔥', 'tipo': 'PESSOAL', 'prioridade': 'ALTA', 'titulo': 'Alerta de Burnout: Superposição de Carga de Trabalho',
                                    'analise': f"No mês selecionado, o consultor {cons_critico['Consultor']} registrou {cons_critico['Hrs_Real']:.0f} horas, um volume insustentável que eleva o risco de burnout, queda de qualidade e turnover.",
                                    'prescricao': "1. Ação Imediata: Revisar a alocação de {cons_critico['Consultor']} para redistribuir tarefas. \n2. Observação Importante: Esta análise se baseia unicamente no volume de horas. Uma avaliação completa de burnout deve considerar fatores qualitativos e pessoais, que estão além do escopo destes dados."})

        if not prescricoes:
             prescricoes.append({'icone': '✅', 'tipo': 'SUCESSO', 'prioridade': 'BAIXA', 'titulo': 'Sinfonia em Harmonia: Excelência Operacional',
                                'analise': "Nossa análise quântica não detectou dissonâncias ou riscos críticos nos dados atuais. Todos os indicadores sistêmicos estão em ressonância positiva.",
                                'prescricao': "1. Documentar Boas Práticas: Identificar os padrões dos projetos de sucesso e transformá-los em metodologia. \n2. Reconhecimento: Celebrar os resultados com a equipe para manter o moral elevado."})
        return prescricoes

    def calcular_metricas_consolidadas(self):
        df = self.estado_quantum
        if df.empty: return {'receita': 0, 'custo': 0, 'lucro': 0, 'margem': 0, 'projetos': 0, 'consultores': 0, 'atraso_medio': 0}
        return {'receita': df['Receita'].sum(), 'custo': df['Custo'].sum(), 'lucro': df['Lucro'].sum(),
                'margem': df['Margem'].mean(), 'projetos': df['Projeto'].nunique(), 'consultores': df['Consultor'].nunique(),
                'atraso_medio': df['Atraso_Dias'].mean()}

# --- FUNÇÃO DE EXPORTAÇÃO EXCEL REFINADA ---
def to_excel_formatted(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Fechamento')
        workbook = writer.book
        worksheet = writer.sheets['Fechamento']
        # Formatação
        for col_idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e2a52", end_color="1e2a52", fill_type="solid")
            
            if 'Total' in col or 'Gerada' in col:
                for row_idx in range(2, len(df) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).number_format = 'R$ #,##0.00'
            elif 'Margem' in col:
                for row_idx in range(2, len(df) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).number_format = '0.0"%"'
    return output.getvalue()


# --- INTERFACE STREAMLIT (VERSÃO DIAMANTE) ---
if 'crq' not in st.session_state:
    st.session_state.crq = CoreQuantumReasoning()
    st.session_state.crq.carregar_universo_dados()
crq = st.session_state.crq

st.markdown('<div class="header-premium"><div class="logo-maestro">🔮 MAESTRO FAROL</div><div class="tagline">A Orquestra de Realidades para a Gestão de Negócios</div></div>', unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 🧭 Controles do Universo")
    mes_sel = st.selectbox("📅 Mês", ["TODOS"] + list(range(1, 13)), index=0)
    ano_sel = st.selectbox("📆 Ano", ["TODOS"] + sorted(crq.dados_universo['Ano'].unique().astype(int)), index=0)
    cons_sel = st.multiselect("👥 Consultores", ["TODOS"] + consultores, default=["TODOS"])
    cli_sel = st.multiselect("🏢 Clientes", ["TODOS"] + clientes, default=["TODOS"])
    tipo_sel = st.multiselect("🎯 Tipo de Serviço", ["TODOS"] + tipos_proj, default=["TODOS"])
    st.button("Executar Análise Quântica", type="primary", use_container_width=True)

filtros = {'consultores': cons_sel, 'clientes': cli_sel, 'tipos': tipo_sel, 'mes': mes_sel, 'ano': ano_sel}
df_filtrado = crq.aplicar_colapso_quantico(filtros)
metricas = crq.calcular_metricas_consolidadas()
indice_saude, status_saude, desc_saude = crq.calcular_indice_saude_quantica()
prescricoes = crq.gerar_prescricoes_quantum_premium(filtros)

tab_executiva, tab_ressonancia, tab_dimensional, tab_fechamento, tab_comparativo, tab_simulador = st.tabs([
    "🎯 Visão Executiva", "🧠 Ressonância Preditiva", "📊 Análises Dimensionais", "💰 Fechamento", "⚖️ Comparativo", "🔮 Simulador Quântico"
])

with tab_executiva:
    st.markdown('<div class="ceo-dashboard">', unsafe_allow_html=True)
    col_saude, col_kpis = st.columns([1, 2])
    with col_saude:
        status_class = f"status-{status_saude.lower()}"
        st.markdown(f'<div class="quantum-health-card {status_class}"><div class="quantum-health-title">Saúde Quântica</div><div class="quantum-health-score {status_class}">{indice_saude}</div></div>', unsafe_allow_html=True)
        st.markdown(f"<p style='text-align: center; margin-top: 10px;'><b>{status_saude}</b>: {desc_saude}</p>", unsafe_allow_html=True)
    with col_kpis:
        kpi1, kpi2, kpi3 = st.columns(3); kpi4, kpi5, kpi6 = st.columns(3)
        kpi1.metric("💰 Receita Total", f"R$ {metricas['receita']:,.0f}"); kpi2.metric("📈 Lucro Total", f"R$ {metricas['lucro']:,.0f}"); kpi3.metric("📊 Margem Média", f"{metricas['margem']:.1f}%")
        kpi4.metric("📁 Projetos Ativos", f"{metricas['projetos']}"); kpi5.metric("👥 Consultores Ativos", f"{metricas['consultores']}"); kpi6.metric("⏳ Atraso Médio (dias)", f"{metricas['atraso_medio']:.1f}")
    st.markdown("<hr style='border-color: rgba(79, 195, 247, 0.1);'>", unsafe_allow_html=True)
    col_mapa, col_receita = st.columns(2)
    with col_mapa:
        st.markdown("#### 🌍 Mapa de Entrelaçamentos")
        mapa_fig = crq.gerar_mapa_entrelacamento()
        if mapa_fig: st.plotly_chart(mapa_fig, use_container_width=True)
        else: st.info("Dados insuficientes para gerar o mapa.")
    with col_receita:
        st.markdown("#### 💰 Receita por Cliente")
        if not df_filtrado.empty:
            fig_rec = px.bar(df_filtrado, x='Cliente', y='Receita', title='', color='TipoProj', template='plotly_dark')
            fig_rec.update_layout(showlegend=False, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig_rec, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

with tab_ressonancia:
    st.markdown("## 🧠 Ressonância Preditiva")
    st.info("O Maestro regeu a sinfonia dos dados e estas são as ressonâncias que emergiram.")
    if prescricoes:
        for p in prescricoes:
            st.markdown(f'<div class="prescription-card {p["prioridade"]}'f'"><h4 class="prescription-title"><span class="prescription-icon">{p["icone"]}</span> {p["titulo"]}</h4><p><b>Análise do Maestro:</b> {p["analise"]}</p><p><b>Prescrição para Ressonância:</b><br>{p["prescricao"]}</p><span style="font-size: 0.8em; color: #aaa; float: right;">TIPO: {p["tipo"]} | PRIORIDADE: {p["prioridade"]}</span></div>', unsafe_allow_html=True)

with tab_dimensional:
    st.markdown("## 📊 Análises Dimensionais")
    if not df_filtrado.empty:
        st.markdown("#### Receita por Cliente por Tipo de Serviço")
        fig_rec_serv = px.bar(df_filtrado, x="Cliente", y="Receita", color="TipoProj", barmode="group", template='plotly_dark')
        fig_rec_serv.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_rec_serv, use_container_width=True)
        
        st.markdown("#### Horas Previstas vs. Realizadas por Consultor")
        horas_consultor = df_filtrado.groupby('Consultor')[['Hrs_Prev', 'Hrs_Real']].sum().reset_index()
        fig_horas = go.Figure(data=[
            go.Bar(name='Previstas', x=horas_consultor['Consultor'], y=horas_consultor['Hrs_Prev']),
            go.Bar(name='Realizadas', x=horas_consultor['Consultor'], y=horas_consultor['Hrs_Real'])
        ])
        fig_horas.update_layout(barmode='group', template='plotly_dark', paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_horas, use_container_width=True)
    else:
        st.warning("Filtre dados para visualizar as análises.")

with tab_fechamento:
    st.markdown("## 💰 Fechamento de Alta Precisão")
    if not df_filtrado.empty:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("### 👥 Por Consultor")
            fech_consultor = df_filtrado.groupby('Consultor').agg(Hrs_Prev=('Hrs_Prev', 'sum'), Hrs_Real=('Hrs_Real', 'sum'), Custo_Total=('Custo', 'sum'), Receita_Gerada=('Receita', 'sum'), Margem_Media=('Margem', 'mean')).reset_index()
            fech_consultor['Desvio_Hrs'] = fech_consultor['Hrs_Real'] - fech_consultor['Hrs_Prev']
            st.dataframe(fech_consultor[['Consultor', 'Hrs_Prev', 'Hrs_Real', 'Desvio_Hrs', 'Custo_Total', 'Receita_Gerada', 'Margem_Media']], use_container_width=True)
            st.download_button("📥 Exportar Consultores (XLSX)", to_excel_formatted(fech_consultor), "fechamento_consultores.xlsx")
        with col2:
            st.markdown("### 🏢 Por Cliente")
            fech_cliente = df_filtrado.groupby('Cliente').agg(Hrs_Prev=('Hrs_Prev', 'sum'), Hrs_Real=('Hrs_Real', 'sum'), Custo_Total=('Custo', 'sum'), Receita_Total=('Receita', 'sum'), Margem_Media=('Margem', 'mean')).reset_index()
            fech_cliente['Desvio_Hrs'] = fech_cliente['Hrs_Real'] - fech_cliente['Hrs_Prev']
            st.dataframe(fech_cliente[['Cliente', 'Hrs_Prev', 'Hrs_Real', 'Desvio_Hrs', 'Custo_Total', 'Receita_Total', 'Margem_Media']], use_container_width=True)
            st.download_button("📥 Exportar Clientes (XLSX)", to_excel_formatted(fech_cliente), "fechamento_clientes.xlsx")
    else:
        st.warning("Nenhum dado para o fechamento com os filtros atuais.")

with tab_comparativo:
    st.markdown("## ⚖️ Comparativo de Realidades")
    hoje = datetime.now()
    ano_atual, mes_atual = hoje.year, hoje.month
    mes_passado = (hoje.replace(day=1) - timedelta(days=1)).month
    ano_mes_passado = (hoje.replace(day=1) - timedelta(days=1)).year
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("##### Realidade 1")
        ano1 = st.selectbox("Ano 1", sorted(crq.dados_universo['Ano'].unique().astype(int)), key='ano1', index=sorted(crq.dados_universo['Ano'].unique().astype(int)).index(ano_mes_passado))
        mes1 = st.selectbox("Mês 1", list(range(1, 13)), key='mes1', index=mes_passado-1)
    with col2:
        st.markdown("##### Realidade 2")
        ano2 = st.selectbox("Ano 2", sorted(crq.dados_universo['Ano'].unique().astype(int)), key='ano2', index=sorted(crq.dados_universo['Ano'].unique().astype(int)).index(ano_atual))
        mes2 = st.selectbox("Mês 2", list(range(1, 13)), key='mes2', index=mes_atual-1)
        
    if st.button("Comparar Realidades", use_container_width=True):
        df1 = crq.dados_universo[(crq.dados_universo['Ano'] == ano1) & (crq.dados_universo['Mes'] == mes1)]
        df2 = crq.dados_universo[(crq.dados_universo['Ano'] == ano2) & (crq.dados_universo['Mes'] == mes2)]
        
        if df1.empty or df2.empty: st.error("Um ou ambos os períodos não contêm dados.")
        else:
            m1 = {'receita': df1['Receita'].sum(), 'lucro': df1['Lucro'].sum(), 'margem': df1['Margem'].mean()}
            m2 = {'receita': df2['Receita'].sum(), 'lucro': df2['Lucro'].sum(), 'margem': df2['Margem'].mean()}
            
            c1,c2,c3 = st.columns(3)
            c1.metric("Receita", f"R$ {m2['receita']:,.0f}", f"R$ {m2['receita'] - m1['receita']:,.0f}")
            c2.metric("Lucro", f"R$ {m2['lucro']:,.0f}", f"R$ {m2['lucro'] - m1['lucro']:,.0f}")
            c3.metric("Margem", f"{m2['margem']:.1f}%", f"{m2['margem'] - m1['margem']:.1f} pp")

            # Análise Detalhada da Variação
            df_comp = pd.concat([df1.assign(Periodo='P1'), df2.assign(Periodo='P2')])
            lucro_proj_comp = df_comp.pivot_table(index='Projeto', columns='Periodo', values='Lucro', aggfunc='sum').fillna(0)
            lucro_proj_comp['Variacao'] = lucro_proj_comp['P2'] - lucro_proj_comp['P1']
            
            top_positivos = lucro_proj_comp[lucro_proj_comp['Variacao'] > 0].sort_values('Variacao', ascending=False).head(3)
            top_negativos = lucro_proj_comp[lucro_proj_comp['Variacao'] < 0].sort_values('Variacao', ascending=True).head(3)

            st.markdown("---")
            st.markdown("#### Vetores de Variação de Lucro")
            v1, v2 = st.columns(2)
            with v1: st.write("📈 **Principais Aumentos:**"); st.dataframe(top_positivos)
            with v2: st.write("📉 **Principais Reduções:**"); st.dataframe(top_negativos)

with tab_simulador:
    st.markdown("## 🔮 Simulador de Realidades")
    st.markdown('<div class="simulator-section">', unsafe_allow_html=True)
    st.write("Aqui, você não prevê o futuro. Você o projeta.")
    
    sim_tipo = st.radio("Escolha o tipo de simulação:", ('Realocação Estratégica', 'Variação de Custo'), horizontal=True, label_visibility="collapsed")
    alteracoes_sim = {}
    
    if sim_tipo == 'Variação de Custo':
        variacao_custo = st.slider("Variação % no Custo/Hora de todos os consultores", -20.0, 20.0, 0.0, 0.5)
        alteracoes_sim = {'tipo': 'custo', 'variacao': variacao_custo}
    
    elif sim_tipo == 'Realocação Estratégica' and not df_filtrado.empty:
        rentabilidade_tipo = df_filtrado.groupby('TipoProj')['ROI_Hora'].mean()
        if len(rentabilidade_tipo) > 1:
            melhor_tipo = rentabilidade_tipo.idxmax()
            pior_tipo = rentabilidade_tipo.idxmin()
            percentual_realoc = st.slider(f"Realocar % de Horas de '{pior_tipo}' para '{melhor_tipo}'", 0.0, 100.0, 0.0, 1.0)
            alteracoes_sim = {'tipo': 'realocacao', 'pior_negocio': pior_tipo, 'melhor_negocio': melhor_tipo, 'percentual': percentual_realoc}
        else:
            st.warning("Dados insuficientes para simulação de realocação (necessário mais de um tipo de projeto).")

    if st.button("▶️ Simular Interferência", type="primary", use_container_width=True):
        if df_filtrado.empty: st.error("Selecione dados na sidebar antes de simular.")
        else:
            df_simulado, sucesso = crq.simular_cenario(df_filtrado, alteracoes_sim)
            if not sucesso: st.error("Erro na simulação. Verifique os dados de entrada.")
            else:
                lucro_orig = df_filtrado['Lucro'].sum(); margem_orig = df_filtrado['Margem'].mean(); atraso_orig = df_filtrado['Atraso_Dias'].mean()
                lucro_sim = df_simulado['Lucro'].sum(); margem_sim = df_simulado['Margem'].mean(); atraso_sim = df_simulado['Atraso_Dias'].mean()
                st.markdown('<div class="simulator-results">', unsafe_allow_html=True)
                st.markdown("<h5>Resultados do Colapso da Simulação:</h5>", unsafe_allow_html=True)
                res1, res2, res3 = st.columns(3)
                res1.metric("Lucro Total", f"R$ {lucro_sim:,.0f}", f"R$ {lucro_sim - lucro_orig:,.0f}")
                res2.metric("Margem Média", f"{margem_sim:.1f}%", f"{margem_sim - margem_orig:.1f} pp")
                res3.metric("Atraso Médio", f"{atraso_sim:.1f} dias", f"{atraso_sim - atraso_orig:.1f} d")
                st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
